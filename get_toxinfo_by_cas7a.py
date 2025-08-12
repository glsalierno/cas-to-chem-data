import aiohttp
import asyncio
import json
from typing: List, Dict, Any
import sys
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_URL = "https://pubchem.ncbi.nlm.nih.gov/rest/pug"
VIEW_URL = "https://pubchem.ncbi.nlm.nih.gov/rest/pug_view"

async def fetch_url(session: aiohttp.ClientSession, url: str, retries: int = 3) -> Dict[str, Any]:
    for attempt in range(retries):
        try:
            async with session.get(url) as response:
                if response.status == 200:
                    return await response.json(content_type=None)
                elif response.status == 503:
                    await asyncio.sleep(2 ** attempt)
                else:
                    print(f"Error {response.status} for URL: {url}")
                    return {'error': f"HTTP error {response.status}"}
        except Exception as e:
            print(f"Exception for URL {url}: {str(e)}")
            if attempt == retries - 1:
                return {'error': str(e)}
            await asyncio.sleep(2 ** attempt)

async def get_pubchem_cid(session: aiohttp.ClientSession, cas_number: str) -> int:
    url = f"{BASE_URL}/compound/name/{quote(cas_number)}/cids/JSON"
    data = await fetch_url(session, url)
    if 'IdentifierList' in data and 'CID' in data['IdentifierList']:
        return data['IdentifierList']['CID'][0]
    print(f"Failed to get CID for CAS {cas_number}. Response: {data}")
    return None

async def get_compound_data(session: aiohttp.ClientSession, cid: int) -> Dict[str, Any]:
    url = f"{VIEW_URL}/data/compound/{cid}/JSON"
    return await fetch_url(session, url)

async def get_iupac_and_smiles(session: aiohttp.ClientSession, cid: int) -> Dict[str, str]:
    url = f"{BASE_URL}/compound/cid/{cid}/property/IUPACName,CanonicalSMILES/JSON"
    data = await fetch_url(session, url)
    if 'PropertyTable' in data and 'Properties' in data['PropertyTable']:
        props = data['PropertyTable']['Properties'][0]
        return {
            'IUPAC': props.get('IUPACName', 'N/A'),
            'SMILES': props.get('CanonicalSMILES', 'N/A')
        }
    return {'IUPAC': 'N/A', 'SMILES': 'N/A'}

def extract_tox_data(compound_data: Dict[str, Any]) -> Dict[str, List[str]]:
    tox_data = {}
    if 'Record' in compound_data and 'Section' in compound_data['Record']:
        for section in compound_data['Record']['Section']:
            if 'Section' in section:
                for sub_section in section['Section']:
                    if 'TOCHeading' in sub_section and 'toxic' in sub_section['TOCHeading'].lower() or 'safety' in sub_section['TOCHeading'].lower() or 'hazard' in sub_section['TOCHeading'].lower():
                        category = sub_section['TOCHeading']
                        if category not in tox_data:
                            tox_data[category] = []
                        if 'Information' in sub_section:
                            for info in sub_section['Information']:
                                if 'Value' in info:
                                    if 'StringWithMarkup' in info['Value']:
                                        tox_data[category].extend([markup['String'] for markup in info['Value']['StringWithMarkup']])
                                    elif 'Number' in info['Value']:
                                        tox_data[category].append(str(info['Value']['Number']))
    return tox_data

async def get_tox_data_for_cas(session: aiohttp.ClientSession, cas_number: str) -> Dict[str, Any]:
    cid = await get_pubchem_cid(session, cas_number)
    if cid is None:
        return {'CAS': cas_number, 'error': 'CID not found'}
    
    compound_data = await get_compound_data(session, cid)
    if 'error' in compound_data:
        return {'CAS': cas_number, 'error': compound_data['error']}
    
    iupac_smiles = await get_iupac_and_smiles(session, cid)
    
    tox_data = extract_tox_data(compound_data)
    
    result = {
        'CAS': cas_number,
        'PubChemCID': cid,
        'IUPAC': iupac_smiles['IUPAC'],
        'SMILES': iupac_smiles['SMILES'],
        'ToxData': tox_data
    }
    
    # Add more fields if needed, like names, synonyms, etc.
    
    return result

async def get_tox_data_for_cas_numbers(cas_numbers: List[str]) -> List[Dict[str, Any]]:
    async with aiohttp.ClientSession() as session:
        sem = asyncio.Semaphore(5)  # Limit to 5 concurrent requests to respect rate limits
        
        async def limited_get(cas):
            async with sem:
                return await get_tox_data_for_cas(session, cas)
        
        tasks = [limited_get(cas) for cas in cas_numbers]
        return await asyncio.gather(*tasks, return_exceptions=True)

def save_to_json(results: List[Dict[str, Any]], filename: str):
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=4)

def save_to_excel(results: List[Dict[str, Any]], filename: str):
    wb = Workbook()
    for compound in results:
        ws = wb.create_sheet(title=compound['CAS'][:31])  # Sheet names limited to 31 chars
        bold_font = Font(bold=True)
        
        row = 1
        ws.cell(row=row, column=1, value="CAS").font = bold_font
        ws.cell(row=row, column=2, value=compound['CAS'])
        row += 1
        
        ws.cell(row=row, column=1, value="PubChem CID").font = bold_font
        ws.cell(row=row, column=2, value=compound.get('PubChemCID', 'N/A'))
        row += 1
        
        ws.cell(row=row, column=1, value="IUPAC Name").font = bold_font
        ws.cell(row=row, column=2, value=compound.get('IUPAC', 'N/A'))
        row += 1
        
        ws.cell(row=row, column=1, value="SMILES").font = bold_font
        ws.cell(row=row, column=2, value=compound.get('SMILES', 'N/A'))
        row += 2
        
        if 'error' in compound:
            ws.cell(row=row, column=1, value=f"Error: {compound['error']}")
        elif 'ToxData' in compound:
            for category, data in compound['ToxData'].items():
                ws.cell(row=row, column=1, value=category).font = bold_font
                row += 1
                for item in data:
                    ws.cell(row=row, column=2, value=item)
                    row += 1
                row += 1
        else:
            ws.cell(row=row, column=1, value="No toxicological data found")

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.remove(wb['Sheet'])  # Remove default empty sheet
    wb.save(filename)
    print(f"Data exported to Excel file: {filename}")

def get_cas_numbers_from_user() -> List[str]:
    print("Enter CAS numbers (one per line). Press Enter twice to finish:")
    cas_numbers = []
    while True:
        cas = input().strip()
        if cas:
            cas_numbers.append(cas)
        else:
            break
    return cas_numbers

if __name__ == "__main__":
    if sys.platform.startswith('win'):
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    if len(sys.argv) > 1:
        cas_numbers = sys.argv[1:]
    else:
        cas_numbers = get_cas_numbers_from_user()

    if not cas_numbers:
        print("No CAS numbers provided. Exiting.")
        sys.exit(1)

    print(f"Fetching data for {len(cas_numbers)} CAS numbers...")
    
    async def main():
        results = await get_tox_data_for_cas_numbers(cas_numbers)
        
        json_filename = "tox_data.json"
        save_to_json(results, json_filename)
        print(f"Data for {len(results)} compounds saved to {json_filename}")

        excel_filename = "tox_data.xlsx"
        save_to_excel(results, excel_filename)

        print("Please check the JSON and Excel files for the retrieved data.")

    asyncio.run(main())
